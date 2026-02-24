"""
Financial Analysis Web App — Backend (Flask) v3
Features: Interactive chart overlays, always-fetch financials, indicator tooltips, competitor analysis
"""
from flask import Flask, render_template, request, jsonify, send_file, Response
from financetoolkit import Toolkit
import requests as http_requests  # for Ollama + external API calls
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import feedparser
import pandas as pd
import numpy as np
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta
from difflib import SequenceMatcher
import json, traceback, time, math

app = Flask(__name__)

class SafeJSONEncoder(json.JSONEncoder):
    """JSON encoder that handles NaN, Inf, numpy types, pandas types."""
    def default(self, obj):
        if isinstance(obj, (np.integer,)): return int(obj)
        if isinstance(obj, (np.floating,)):
            f = float(obj)
            if math.isnan(f) or math.isinf(f): return None
            return f
        if isinstance(obj, (np.bool_,)): return bool(obj)
        if isinstance(obj, np.ndarray): return obj.tolist()
        if isinstance(obj, (pd.Timestamp, pd.Period)): return str(obj)
        if hasattr(obj, 'item'): return obj.item()
        return super().default(obj)

def safe_jsonify(data, status=200):
    """Serialize to JSON safely, replacing NaN/Inf with null."""
    text = json.dumps(data, cls=SafeJSONEncoder, allow_nan=True)
    # Replace JavaScript-invalid NaN/Infinity tokens with null
    text = text.replace(': NaN', ': null').replace(':NaN', ':null')
    text = text.replace(': Infinity', ': null').replace(':Infinity', ':null')
    text = text.replace(': -Infinity', ': null').replace(':-Infinity', ':null')
    return Response(text, status=status, mimetype='application/json')

FMP_API_KEY = "wybWEsp1oB9abHfz3yPpQYwffxaN21B7"
GEMINI_API_KEY = "AIzaSyBX8v2d_UV_Hktcj-AvV7TyR6TD5grn24w"  # fallback
ALPHA_VANTAGE_KEY = "P0D5N0A8SVC00YUW"
FINNHUB_KEY = "d6ao9dhr01qqjvbr6m1gd6ao9dhr01qqjvbr6m20"
FRED_KEY = "010a35b0ca89efdef4234f33c5089d7a"

# ── AI Backend: Ollama (local) with Gemini fallback ──────
OLLAMA_URL = "http://localhost:11434"
OLLAMA_MODEL = "llama3.1"  # 8B params, runs well on 3050 GPU

def call_ai(prompt):
    """Try Ollama local first, fall back to Gemini API if unavailable."""
    # Try Ollama (local, free, instant)
    try:
        r = http_requests.post(f"{OLLAMA_URL}/api/generate", json={
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": 0.7, "num_predict": 2048}
        }, timeout=120)
        if r.status_code == 200:
            return r.json().get("response", "")
    except Exception as e:
        print(f"Ollama unavailable: {e}")

    # Fallback: Gemini API
    try:
        from google import genai
        client = genai.Client(api_key=GEMINI_API_KEY)
        response = client.models.generate_content(model="gemini-2.0-flash", contents=prompt)
        return response.text
    except Exception as e:
        print(f"Gemini fallback also failed: {e}")
        return None

# ── TICKER DATABASE ──────────────────────────────────────
TICKER_DB = [
    ("AAPL","Apple"),("ABBV","AbbVie"),("ABT","Abbott Labs"),("ACN","Accenture"),
    ("ADBE","Adobe"),("ADI","Analog Devices"),("ADM","Archer-Daniels"),("ADP","ADP"),
    ("ADSK","Autodesk"),("AEP","American Electric"),("AFL","Aflac"),("AIG","AIG"),
    ("AMAT","Applied Materials"),("AMD","AMD"),("AMGN","Amgen"),("AMP","Ameriprise"),
    ("AMZN","Amazon"),("ANET","Arista Networks"),("ANSS","Ansys"),("AON","Aon"),
    ("APD","Air Products"),("APH","Amphenol"),("AVGO","Broadcom"),("AXP","American Express"),
    ("BA","Boeing"),("BAC","Bank of America"),("BAX","Baxter"),("BDX","Becton Dickinson"),
    ("BK","Bank of NY Mellon"),("BKNG","Booking Holdings"),("BLK","BlackRock"),
    ("BMY","Bristol-Myers"),("BRK.B","Berkshire Hathaway"),("BSX","Boston Scientific"),
    ("C","Citigroup"),("CAT","Caterpillar"),("CB","Chubb"),("CCI","Crown Castle"),
    ("CDNS","Cadence Design"),("CI","Cigna"),("CL","Colgate-Palmolive"),("CMCSA","Comcast"),
    ("CME","CME Group"),("CNC","Centene"),("COF","Capital One"),("COP","ConocoPhillips"),
    ("COST","Costco"),("CRM","Salesforce"),("CSCO","Cisco"),("CTAS","Cintas"),
    ("CVS","CVS Health"),("CVX","Chevron"),("D","Dominion Energy"),("DD","DuPont"),
    ("DE","Deere & Co"),("DHR","Danaher"),("DIS","Disney"),("DLTR","Dollar Tree"),
    ("DOW","Dow Inc"),("DUK","Duke Energy"),("DVN","Devon Energy"),("DXCM","DexCom"),
    ("EA","Electronic Arts"),("EBAY","eBay"),("ECL","Ecolab"),("EL","Estee Lauder"),
    ("EMR","Emerson"),("ENPH","Enphase Energy"),("EOG","EOG Resources"),("EQIX","Equinix"),
    ("EW","Edwards Lifesciences"),("EXC","Exelon"),("F","Ford"),("FAST","Fastenal"),
    ("FCX","Freeport-McMoRan"),("FDX","FedEx"),("FSLR","First Solar"),
    ("GD","General Dynamics"),("GE","GE Aerospace"),("GILD","Gilead"),("GIS","General Mills"),
    ("GM","General Motors"),("GOOG","Alphabet A"),("GOOGL","Alphabet C"),("GPN","Global Payments"),
    ("GS","Goldman Sachs"),("HAL","Halliburton"),("HD","Home Depot"),("HON","Honeywell"),
    ("HPQ","HP Inc"),("HUM","Humana"),("IBM","IBM"),("ICE","Intercontinental Exchange"),
    ("IDXX","IDEXX Labs"),("ILMN","Illumina"),("INTC","Intel"),("INTU","Intuit"),
    ("ISRG","Intuitive Surgical"),("ITW","Illinois Tool Works"),("JCI","Johnson Controls"),
    ("JNJ","Johnson & Johnson"),("JPM","JPMorgan Chase"),("KHC","Kraft Heinz"),
    ("KLAC","KLA Corp"),("KMB","Kimberly-Clark"),("KO","Coca-Cola"),("LEN","Lennar"),
    ("LHX","L3Harris"),("LIN","Linde"),("LLY","Eli Lilly"),("LMT","Lockheed Martin"),
    ("LOW","Lowe's"),("LRCX","Lam Research"),("LULU","Lululemon"),("MA","Mastercard"),
    ("MAR","Marriott"),("MCD","McDonald's"),("MCHP","Microchip Tech"),("MCK","McKesson"),
    ("MCO","Moody's"),("MDLZ","Mondelez"),("MDT","Medtronic"),("MET","MetLife"),
    ("META","Meta Platforms"),("MMC","Marsh McLennan"),("MMM","3M"),("MO","Altria"),
    ("MPC","Marathon Petroleum"),("MRK","Merck"),("MRNA","Moderna"),("MS","Morgan Stanley"),
    ("MSFT","Microsoft"),("MSI","Motorola Solutions"),("MU","Micron"),("NFLX","Netflix"),
    ("NKE","Nike"),("NOC","Northrop Grumman"),("NOW","ServiceNow"),("NSC","Norfolk Southern"),
    ("NTAP","NetApp"),("NVDA","NVIDIA"),("NVO","Novo Nordisk"),("NXPI","NXP Semi"),
    ("O","Realty Income"),("ODFL","Old Dominion"),("ON","ON Semi"),("ORCL","Oracle"),
    ("ORLY","O'Reilly Auto"),("OXY","Occidental"),("PANW","Palo Alto Networks"),
    ("PARA","Paramount"),("PCAR","PACCAR"),("PEP","PepsiCo"),("PFE","Pfizer"),
    ("PG","Procter & Gamble"),("PGR","Progressive"),("PLD","Prologis"),("PLTR","Palantir"),
    ("PM","Philip Morris"),("PNC","PNC Financial"),("PSA","Public Storage"),("PSX","Phillips 66"),
    ("PYPL","PayPal"),("QCOM","Qualcomm"),("REGN","Regeneron"),("RIVN","Rivian"),
    ("ROKU","Roku"),("ROP","Roper Tech"),("ROST","Ross Stores"),("RTX","RTX Corp"),
    ("SBUX","Starbucks"),("SCHW","Charles Schwab"),("SHW","Sherwin-Williams"),
    ("SLB","Schlumberger"),("SMCI","Super Micro"),("SNAP","Snap Inc"),("SNPS","Synopsys"),
    ("SO","Southern Co"),("SOFI","SoFi"),("SPG","Simon Property"),("SPGI","S&P Global"),
    ("SQ","Block Inc"),("SRE","Sempra"),("STZ","Constellation Brands"),("SYK","Stryker"),
    ("SYY","Sysco"),("T","AT&T"),("TDG","TransDigm"),("TGT","Target"),("TJX","TJX Cos"),
    ("TMO","Thermo Fisher"),("TMUS","T-Mobile"),("TSLA","Tesla"),("TSM","TSMC"),
    ("TSN","Tyson Foods"),("TXN","Texas Instruments"),("UNH","UnitedHealth"),
    ("UNP","Union Pacific"),("UPS","UPS"),("URI","United Rentals"),("USB","US Bancorp"),
    ("V","Visa"),("VLO","Valero Energy"),("VRSK","Verisk"),("VRTX","Vertex Pharma"),
    ("VZ","Verizon"),("WBA","Walgreens"),("WBD","Warner Bros Discovery"),("WELL","Welltower"),
    ("WFC","Wells Fargo"),("WM","Waste Management"),("WMT","Walmart"),("XEL","Xcel Energy"),
    ("XOM","ExxonMobil"),("ZM","Zoom"),("ZTS","Zoetis"),
    ("RELIANCE.NS","Reliance Industries"),("TCS.NS","TCS"),("INFY.NS","Infosys"),
    ("HDFCBANK.NS","HDFC Bank"),("ICICIBANK.NS","ICICI Bank"),("SBIN.NS","SBI"),
    ("BHARTIARTL.NS","Bharti Airtel"),("ITC.NS","ITC"),("KOTAKBANK.NS","Kotak Bank"),
    ("LT.NS","Larsen & Toubro"),("HINDUNILVR.NS","Hindustan Unilever"),
    ("BAJFINANCE.NS","Bajaj Finance"),("MARUTI.NS","Maruti Suzuki"),("WIPRO.NS","Wipro"),
    ("TATAMOTORS.NS","Tata Motors"),("TATASTEEL.NS","Tata Steel"),("SUNPHARMA.NS","Sun Pharma"),
    ("AXISBANK.NS","Axis Bank"),("ONGC.NS","ONGC"),("NTPC.NS","NTPC"),
    ("ADANIENT.NS","Adani Enterprises"),("TITAN.NS","Titan"),("ASIANPAINT.NS","Asian Paints"),
    # ── Futures ──
    ("ES=F","S&P 500 Futures"),("NQ=F","Nasdaq 100 Futures"),("YM=F","Dow Futures"),
    ("RTY=F","Russell 2000 Futures"),("GC=F","Gold Futures"),("SI=F","Silver Futures"),
    ("CL=F","Crude Oil WTI Futures"),("BZ=F","Brent Crude Futures"),("NG=F","Natural Gas Futures"),
    ("ZC=F","Corn Futures"),("ZW=F","Wheat Futures"),("ZS=F","Soybean Futures"),
    ("HG=F","Copper Futures"),("PL=F","Platinum Futures"),("PA=F","Palladium Futures"),
    ("ZB=F","US Treasury Bond Futures"),("ZN=F","10-Year Note Futures"),
    # ── Currency Pairs ──
    ("EURUSD=X","EUR/USD"),("GBPUSD=X","GBP/USD"),("USDJPY=X","USD/JPY"),
    ("USDCHF=X","USD/CHF"),("AUDUSD=X","AUD/USD"),("USDCAD=X","USD/CAD"),
    ("NZDUSD=X","NZD/USD"),("USDINR=X","USD/INR"),("GBPINR=X","GBP/INR"),
    ("EURINR=X","EUR/INR"),("USDHKD=X","USD/HKD"),("USDSGD=X","USD/SGD"),
    ("EURGBP=X","EUR/GBP"),("EURJPY=X","EUR/JPY"),("GBPJPY=X","GBP/JPY"),
    # ── Indices (ETFs) ──
    ("SPY","SPDR S&P 500 ETF"),("QQQ","Invesco Nasdaq 100 ETF"),("DIA","SPDR Dow Jones ETF"),
    ("IWM","iShares Russell 2000"),("VTI","Vanguard Total Market"),("EEM","iShares EM ETF"),
    ("GLD","SPDR Gold Shares"),("SLV","iShares Silver Trust"),("TLT","iShares 20+ Yr Bond"),
    ("^NSEI","NIFTY 50"),("^BSESN","BSE SENSEX"),
    # ── Global Indices ──
    ("^GSPC","S&P 500"),("^IXIC","NASDAQ Composite"),("^DJI","Dow Jones Industrial"),
    ("^FTSE","FTSE 100"),("^N225","Nikkei 225"),("^GDAXI","DAX"),("^HSI","Hang Seng"),
    ("000001.SS","Shanghai Composite"),("^AXJO","ASX 200"),("^FCHI","CAC 40"),
    ("^RUT","Russell 2000"),("^VIX","CBOE Volatility Index"),
]

# ── SECTOR PEER GROUPS ───────────────────────────────────
SECTOR_PEERS = {
    "tech": ["AAPL","MSFT","GOOGL","META","AMZN","NVDA","TSLA","AMD","INTC","CRM","ADBE","ORCL","IBM","NOW","AVGO"],
    "finance": ["JPM","BAC","GS","MS","WFC","C","BLK","SCHW","AXP","COF","PNC","USB","BK","MCO","SPGI"],
    "healthcare": ["JNJ","UNH","PFE","MRK","ABBV","LLY","TMO","ABT","BMY","AMGN","GILD","MDT","ISRG","REGN","VRTX"],
    "consumer": ["PG","KO","PEP","COST","WMT","HD","MCD","NKE","SBUX","TGT","CL","EL","KHC","MDLZ","GIS"],
    "energy": ["XOM","CVX","COP","EOG","OXY","SLB","HAL","DVN","MPC","VLO","PSX"],
    "industrial": ["CAT","DE","HON","BA","GE","RTX","LMT","NOC","GD","EMR","ITW","UNP","UPS","FDX"],
    "telecom": ["T","VZ","TMUS","CMCSA"],
}

def get_peers(ticker):
    t = ticker.upper()
    for sector, tickers in SECTOR_PEERS.items():
        if t in tickers:
            return [p for p in tickers if p != t][:5], sector
    return [], "unknown"

def safe_val(val):
    if val is None: return None
    if isinstance(val, (np.integer,)): return int(val)
    if isinstance(val, (np.floating,)): return float(val)
    if isinstance(val, (np.bool_,)): return bool(val)
    if isinstance(val, pd.Timestamp): return str(val)
    if isinstance(val, (pd.Series, pd.DataFrame)):
        return safe_val(val.iloc[-1] if len(val) > 0 else 0)
    if hasattr(val, 'item'): return val.item()
    return val

def safe_num(val, decimals=2):
    v = safe_val(val)
    if v is None or (isinstance(v, float) and (np.isnan(v) or np.isinf(v))):
        return "N/A"
    try: return round(float(v), decimals)
    except: return "N/A"

def sanitize_for_json(obj):
    """Recursively clean NaN/Inf/numpy types so jsonify won't crash on Windows."""
    if isinstance(obj, dict):
        return {k: sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [sanitize_for_json(v) for v in obj]
    if isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj): return None
        return obj
    if isinstance(obj, (np.integer,)): return int(obj)
    if isinstance(obj, (np.floating,)):
        f = float(obj)
        return None if np.isnan(f) or np.isinf(f) else f
    if isinstance(obj, (np.bool_,)): return bool(obj)
    if isinstance(obj, (pd.Timestamp, pd.Period)): return str(obj)
    if hasattr(obj, 'item'): return obj.item()
    return obj

def get_fin_val(df, ticker, row_label):
    try:
        if isinstance(df.columns, pd.MultiIndex):
            if ticker in df.columns.get_level_values(1):
                df = df.xs(ticker, level=1, axis=1)
            elif ticker in df.columns.get_level_values(0):
                df = df.xs(ticker, level=0, axis=1)
        if row_label in df.index:
            return safe_num(df.loc[row_label].iloc[-1])
    except: pass
    return "N/A"

# ── Route: Dashboard ─────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

# ── Route: Ticker Suggestions ────────────────────────────
@app.route("/api/suggest", methods=["GET"])
def suggest():
    q = request.args.get("q", "").upper().strip()
    asset_type = request.args.get("asset_type", "").lower().strip()
    if len(q) < 1: return safe_jsonify([])

    def matches_asset_type(ticker):
        if not asset_type: return True
        if asset_type == "stocks":
            return not ticker.endswith("=F") and not ticker.endswith("=X")
        elif asset_type == "futures":
            return ticker.endswith("=F")
        elif asset_type == "currencies":
            return ticker.endswith("=X")
        elif asset_type == "options":
            return not ticker.endswith("=F") and not ticker.endswith("=X") and not ticker.startswith("^")
        return True

    results = []
    for ticker, name in TICKER_DB:
        if not matches_asset_type(ticker): continue
        score = 0
        if ticker.upper().startswith(q): score = 100 + (10 - len(ticker))
        elif q in ticker.upper(): score = 80
        elif q in name.upper(): score = 70
        else:
            ratio = SequenceMatcher(None, q, ticker.upper()).ratio()
            name_ratio = SequenceMatcher(None, q, name.upper()).ratio()
            best = max(ratio, name_ratio)
            if best > 0.5: score = int(best * 60)
        if score > 0:
            results.append({"ticker": ticker, "name": name, "score": score})
    results.sort(key=lambda x: -x["score"])
    return safe_jsonify(results[:10])

# ── Route: Full Analysis ─────────────────────────────────
@app.route("/api/analyze", methods=["POST"])
def analyze():
    data = request.json
    ticker = data.get("ticker", "AAPL").upper().strip()
    period = data.get("period", "yearly")
    start_date = data.get("start_date", "")
    end_date = data.get("end_date", "")

    timeframe = data.get("timeframe", "1Y")
    if not start_date:
        offsets = {"1M": 30, "3M": 90, "6M": 180, "1Y": 365, "2Y": 730, "5Y": 1825, "MAX": 7300}
        days = offsets.get(timeframe, 365)
        start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")

    try:
        # Chart data uses the user's timeframe
        hist = None
        try:
            chart_toolkit = Toolkit(tickers=[ticker], api_key=FMP_API_KEY, start_date=start_date, end_date=end_date)
            hist = chart_toolkit.get_historical_data(period="daily")
        except OSError as oe:
            print(f"FinanceToolkit OSError (Windows caching): {oe}")
            # Fallback: use yfinance for price history
            import yfinance as yf
            yf_ticker = yf.Ticker(ticker)
            yf_hist = yf_ticker.history(start=start_date, end=end_date)
            if not yf_hist.empty:
                yf_hist.index.name = 'date'
                yf_hist.columns = pd.MultiIndex.from_product([yf_hist.columns, [ticker]])
                hist = yf_hist
        except Exception as e:
            print(f"FinanceToolkit error: {e}")
            # Fallback: use yfinance
            import yfinance as yf
            yf_ticker = yf.Ticker(ticker)
            yf_hist = yf_ticker.history(start=start_date, end=end_date)
            if not yf_hist.empty:
                yf_hist.index.name = 'date'
                yf_hist.columns = pd.MultiIndex.from_product([yf_hist.columns, [ticker]])
                hist = yf_hist

        if hist is None or hist.empty:
            return safe_jsonify({"error": f"No data found for {ticker}"}), 404

        if isinstance(hist.columns, pd.MultiIndex):
            if ticker in hist.columns.get_level_values(1):
                hd = hist.xs(ticker, level=1, axis=1)
            elif ticker in hist.columns.get_level_values(0):
                hd = hist.xs(ticker, level=0, axis=1)
            else: hd = hist
        else: hd = hist

        close = hd['Close']
        price = safe_num(close.iloc[-1])

        # Compute technicals from price data
        delta = close.diff()
        gain = delta.where(delta > 0, 0).rolling(14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
        rs = gain / loss
        rsi = safe_num(100 - (100 / (1 + rs)).iloc[-1])

        ema5 = safe_num(close.ewm(span=5).mean().iloc[-1])
        ema10 = safe_num(close.ewm(span=10).mean().iloc[-1])
        ema20 = safe_num(close.ewm(span=20).mean().iloc[-1])
        ema50 = safe_num(close.ewm(span=50).mean().iloc[-1])
        sma200 = safe_num(close.rolling(200).mean().iloc[-1])

        ema12 = close.ewm(span=12).mean()
        ema26 = close.ewm(span=26).mean()
        macd_line = ema12 - ema26
        signal_line = macd_line.ewm(span=9).mean()
        macd_val = safe_num(macd_line.iloc[-1])
        macd_signal = safe_num(signal_line.iloc[-1])

        ma20 = close.rolling(20).mean()
        std20 = close.rolling(20).std()
        bb_upper = safe_num((ma20 + 2 * std20).iloc[-1])
        bb_lower = safe_num((ma20 - 2 * std20).iloc[-1])

        tr = pd.concat([hd['High'] - hd['Low'], (hd['High'] - close.shift()).abs(), (hd['Low'] - close.shift()).abs()], axis=1).max(axis=1)
        atr = safe_num(tr.rolling(14).mean().iloc[-1])

        vol = safe_num(hd['Volume'].iloc[-1])
        avg_vol = safe_num(hd['Volume'].rolling(20).mean().iloc[-1])

        technicals = {
            "price": price, "ema_5": ema5, "ema_10": ema10, "ema_20": ema20,
            "ema_50": ema50, "sma_200": sma200,
            "rsi": rsi, "macd": macd_val, "macd_signal": macd_signal,
            "bb_upper": bb_upper, "bb_lower": bb_lower, "atr": atr,
            "volume": vol, "avg_volume_20": avg_vol
        }

        # ── FINANCIALS — always fetch with 5Y lookback ──
        fin_toolkit = Toolkit(tickers=[ticker], api_key=FMP_API_KEY,
                              quarterly=(period == "quarterly"))
        financials = {}
        try:
            income = fin_toolkit.get_income_statement()
            balance = fin_toolkit.get_balance_sheet_statement()
            financials = {
                "revenue": get_fin_val(income, ticker, "Revenue"),
                "cost_of_goods": get_fin_val(income, ticker, "Cost of Goods Sold"),
                "gross_profit": get_fin_val(income, ticker, "Gross Profit"),
                "operating_income": get_fin_val(income, ticker, "Operating Income"),
                "ebitda": get_fin_val(income, ticker, "EBITDA"),
                "net_income": get_fin_val(income, ticker, "Net Income"),
                "eps": get_fin_val(income, ticker, "EPS"),
                "eps_diluted": get_fin_val(income, ticker, "EPS Diluted"),
                "total_assets": get_fin_val(balance, ticker, "Total Assets"),
                "total_debt": get_fin_val(balance, ticker, "Total Debt"),
                "net_debt": get_fin_val(balance, ticker, "Net Debt"),
                "cash": get_fin_val(balance, ticker, "Cash and Cash Equivalents"),
                "total_equity": get_fin_val(balance, ticker, "Total Equity"),
                "retained_earnings": get_fin_val(balance, ticker, "Retained Earnings"),
            }
        except Exception as e:
            financials = {"error": str(e)}

        # ── RATIOS ──
        ratios = {}
        try:
            prof = fin_toolkit.ratios.collect_profitability_ratios()
            val = fin_toolkit.ratios.collect_valuation_ratios()
            ratios = {
                "gross_margin": get_fin_val(prof, ticker, "Gross Margin"),
                "operating_margin": get_fin_val(prof, ticker, "Operating Margin"),
                "net_profit_margin": get_fin_val(prof, ticker, "Net Profit Margin"),
                "roe": get_fin_val(prof, ticker, "Return on Equity"),
                "roa": get_fin_val(prof, ticker, "Return on Assets"),
                "roic": get_fin_val(prof, ticker, "Return on Invested Capital"),
                "pe_ratio": get_fin_val(val, ticker, "Price-to-Earnings"),
                "pb_ratio": get_fin_val(val, ticker, "Price-to-Book"),
                "ev_ebitda": get_fin_val(val, ticker, "EV-to-EBITDA"),
                "ev_sales": get_fin_val(val, ticker, "EV-to-Sales"),
                "dividend_yield": get_fin_val(val, ticker, "Dividend Yield"),
                "market_cap": get_fin_val(val, ticker, "Market Cap"),
            }
        except Exception as e:
            ratios = {"error": str(e)}

        # ── FCFF ──
        fcff = {}
        try:
            cf = fin_toolkit.get_cash_flow_statement()
            fcff = {
                "cash_flow_from_operations": get_fin_val(cf, ticker, "Cash Flow from Operations"),
                "capital_expenditure": get_fin_val(cf, ticker, "Capital Expenditure"),
                "free_cash_flow": get_fin_val(cf, ticker, "Free Cash Flow"),
                "net_change_in_cash": get_fin_val(cf, ticker, "Net Change in Cash"),
                "cash_flow_from_investing": get_fin_val(cf, ticker, "Cash Flow from Investing"),
                "cash_flow_from_financing": get_fin_val(cf, ticker, "Cash Flow from Financing"),
                "dividends_paid": get_fin_val(cf, ticker, "Dividends Paid"),
                "stock_based_compensation": get_fin_val(cf, ticker, "Stock Based Compensation"),
            }
        except Exception as e:
            fcff = {"error": str(e)}

        # ── Full price history for interactive chart ──
        price_history = []
        try:
            for date, row in hd.iterrows():
                price_history.append({
                    "date": str(date)[:10],
                    "open": safe_num(row.get('Open', 0)),
                    "high": safe_num(row.get('High', 0)),
                    "low": safe_num(row.get('Low', 0)),
                    "close": safe_num(row['Close']),
                    "volume": safe_num(row['Volume'])
                })
        except: pass

        result = sanitize_for_json({
            "ticker": ticker, "technicals": technicals,
            "financials": financials, "ratios": ratios, "fcff": fcff,
            "price_history": price_history,
            "date_range": {"start": start_date, "end": end_date}
        })
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Competitor Analysis ───────────────────────────
def get_multi_val(df, ticker, metric):
    """Extract value from multi-ticker financetoolkit DataFrame.
    Multi-ticker DataFrames have a row MultiIndex: (ticker, metric). Columns are year periods."""
    try:
        if isinstance(df.index, pd.MultiIndex):
            if (ticker, metric) in df.index:
                row = df.loc[(ticker, metric)]
                return safe_num(row.iloc[-1])
        else:
            if metric in df.index:
                return safe_num(df.loc[metric].iloc[-1])
    except: pass
    return "N/A"

@app.route("/api/competitors", methods=["POST"])
def competitors():
    data = request.json
    ticker = data.get("ticker", "AAPL").upper().strip()
    peers, sector = get_peers(ticker)

    if not peers:
        return safe_jsonify({"sector": "unknown", "peers": [], "message": "No peer data available for this ticker."})

    try:
        all_tickers = [ticker] + peers
        comp_toolkit = Toolkit(tickers=all_tickers, api_key=FMP_API_KEY)
        val = comp_toolkit.ratios.collect_valuation_ratios()
        prof = comp_toolkit.ratios.collect_profitability_ratios()

        result = []
        for t in all_tickers:
            entry = {"ticker": t, "is_target": (t == ticker)}
            entry["pe_ratio"] = get_multi_val(val, t, "Price-to-Earnings")
            entry["pb_ratio"] = get_multi_val(val, t, "Price-to-Book")
            entry["ev_ebitda"] = get_multi_val(val, t, "EV-to-EBITDA")
            entry["ev_sales"] = get_multi_val(val, t, "EV-to-Sales")
            entry["market_cap"] = get_multi_val(val, t, "Market Cap")
            entry["net_margin"] = get_multi_val(prof, t, "Net Profit Margin")
            entry["roe"] = get_multi_val(prof, t, "Return on Equity")
            entry["gross_margin"] = get_multi_val(prof, t, "Gross Margin")
            result.append(entry)

        return safe_jsonify({"sector": sector, "peers": result})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"sector": sector, "peers": [], "error": str(e)})

# ── Route: News ──────────────────────────────────────────
@app.route("/api/news", methods=["POST"])
def news():
    data = request.json
    ticker = data.get("ticker", "AAPL").upper().strip()
    try:
        encoded = ticker.replace("&", "%26")
        rss_url = f"https://news.google.com/rss/search?q={encoded}+stock&hl=en-US&gl=US&ceid=US:en"
        feed = feedparser.parse(rss_url)
        analyzer = SentimentIntensityAnalyzer()
        news_items = []
        for entry in feed.entries[:5]:
            title = entry.title
            sentiment = analyzer.polarity_scores(f"{title}. {entry.get('description', '')}")
            news_items.append({
                "title": title, "link": entry.link,
                "published": entry.get("published", ""),
                "sentiment_score": round(sentiment['compound'], 4),
                "sentiment_label": "Positive" if sentiment['compound'] > 0.05 else "Negative" if sentiment['compound'] < -0.05 else "Neutral"
            })
        avg = sum(n['sentiment_score'] for n in news_items) / len(news_items) if news_items else 0
        return safe_jsonify({"ticker": ticker, "news": news_items, "average_sentiment": round(avg, 4),
                        "overall_label": "Positive" if avg > 0.05 else "Negative" if avg < -0.05 else "Neutral"})
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

# ── Route: AI ────────────────────────────────────────────
@app.route("/api/ai", methods=["POST"])
def ai_overview():
    data = request.json
    ticker = data.get("ticker", "")
    analysis_data = data.get("analysis", {})
    news_data = data.get("news", {})
    user_question = data.get("question", "")

    prompt = f"""You are a top-tier financial analyst. Here is the latest data for {ticker}:

TECHNICALS:
{json.dumps(analysis_data.get('technicals', {}), indent=2)}

FINANCIALS:
{json.dumps(analysis_data.get('financials', {}), indent=2)}

RATIOS:
{json.dumps(analysis_data.get('ratios', {}), indent=2)}

FREE CASH FLOW:
{json.dumps(analysis_data.get('fcff', {}), indent=2)}

NEWS SENTIMENT:
{json.dumps(news_data, indent=2)}

Based on the data, provide:
1. Key strengths and competitive advantages
2. Risks and red flags
3. Clear recommendation (Buy / Hold / Sell) with reasoning

{f'The user also asks: {user_question}' if user_question else ''}
"""
    result = call_ai(prompt)
    if result:
        return safe_jsonify({"overview": result})
    return safe_jsonify({"error": "AI unavailable. Please ensure Ollama is running: 'ollama serve' in terminal."}), 500

# ── Route: Excel Export ──────────────────────────────────
@app.route("/api/export", methods=["POST"])
def export_excel():
    data = request.json
    ticker = data.get("ticker", "DATA")
    analysis = data.get("analysis", {})
    news_data = data.get("news", {})
    ai_text = data.get("ai_overview", "")

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Technicals"
    ws1.append(["Indicator", "Value"])
    for k, v in analysis.get("technicals", {}).items():
        ws1.append([k.replace("_"," ").title(), str(v)])

    ws2 = wb.create_sheet("Financials")
    ws2.append(["Metric", "Value"])
    for k, v in analysis.get("financials", {}).items():
        ws2.append([k.replace("_"," ").title(), str(v)])

    ws3 = wb.create_sheet("Ratios & Valuation")
    ws3.append(["Ratio", "Value"])
    for k, v in analysis.get("ratios", {}).items():
        ws3.append([k.replace("_"," ").title(), str(v)])

    ws4 = wb.create_sheet("Cash Flow (FCFF)")
    ws4.append(["Metric", "Value"])
    for k, v in analysis.get("fcff", {}).items():
        ws4.append([k.replace("_"," ").title(), str(v)])

    ws5 = wb.create_sheet("News & Sentiment")
    ws5.append(["Title", "Sentiment", "Score"])
    for n in news_data.get("news", []):
        ws5.append([n.get("title",""), n.get("sentiment_label",""), n.get("sentiment_score","")])

    ws6 = wb.create_sheet("AI Overview")
    ws6.append(["AI Analyst Commentary"])
    for line in ai_text.split("\n"):
        ws6.append([line])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"{ticker}_analysis.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════════════════
# STAGE 2 — NEW ENDPOINTS
# ══════════════════════════════════════════════════════════

# ── Route: DCF Intrinsic Valuation ───────────────────────
@app.route("/api/dcf", methods=["POST"])
def dcf_valuation():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        tk = Toolkit(tickers=[ticker], api_key=FMP_API_KEY)

        income = tk.get_income_statement()
        balance = tk.get_balance_sheet_statement()
        cf = tk.get_cash_flow_statement()

        fcf = get_fin_val(cf, ticker, "Free Cash Flow")
        revenue = get_fin_val(income, ticker, "Revenue")
        ebit = get_fin_val(income, ticker, "Operating Income")
        total_debt = get_fin_val(balance, ticker, "Total Debt")
        cash = get_fin_val(balance, ticker, "Cash and Cash Equivalents")
        total_equity = get_fin_val(balance, ticker, "Total Equity")
        shares = get_fin_val(income, ticker, "Weighted Average Shares Outstanding")
        total_assets = get_fin_val(balance, ticker, "Total Assets")

        if any(v == "N/A" for v in [fcf, total_debt, cash, shares, total_equity]):
            return safe_jsonify({"error": "Insufficient data for DCF calculation", "ticker": ticker})

        # WACC estimation
        risk_free = 0.043  # ~10Y Treasury
        equity_premium = 0.055
        beta = 1.0  # Default
        cost_of_equity = risk_free + beta * equity_premium
        cost_of_debt = 0.05
        tax_rate = 0.21
        total_cap = abs(total_equity) + abs(total_debt) if total_debt != 0 else abs(total_equity)
        eq_weight = abs(total_equity) / total_cap if total_cap > 0 else 0.7
        debt_weight = 1 - eq_weight
        wacc = eq_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate)
        wacc = max(wacc, 0.06)

        # Project FCF
        growth_rate = 0.08   # Conservative growth
        terminal_growth = 0.025
        projection_years = 5
        projected_fcf = []
        current_fcf = float(fcf)
        for y in range(1, projection_years + 1):
            current_fcf *= (1 + growth_rate)
            pv = current_fcf / ((1 + wacc) ** y)
            projected_fcf.append({"year": y, "fcf": round(current_fcf, 0), "pv": round(pv, 0)})

        # Terminal value
        terminal_value = (current_fcf * (1 + terminal_growth)) / (wacc - terminal_growth)
        pv_terminal = terminal_value / ((1 + wacc) ** projection_years)

        # Enterprise value & intrinsic value
        sum_pv_fcf = sum(p["pv"] for p in projected_fcf)
        enterprise_value = sum_pv_fcf + pv_terminal
        equity_value = enterprise_value - float(total_debt) + float(cash)
        intrinsic_per_share = equity_value / float(shares)

        # Get current price for comparison
        import yfinance as yf
        stock = yf.Ticker(ticker)
        current_price = stock.info.get("currentPrice", stock.info.get("previousClose", 0))

        upside = ((intrinsic_per_share - current_price) / current_price * 100) if current_price > 0 else 0
        verdict = "UNDERVALUED" if upside > 15 else "OVERVALUED" if upside < -15 else "FAIRLY VALUED"

        return safe_jsonify(sanitize_for_json({
            "ticker": ticker, "intrinsic_value": round(intrinsic_per_share, 2),
            "current_price": round(current_price, 2), "upside_pct": round(upside, 2),
            "verdict": verdict,
            "wacc": round(wacc * 100, 2), "growth_rate": growth_rate * 100,
            "terminal_growth": terminal_growth * 100,
            "enterprise_value": round(enterprise_value, 0),
            "projected_fcf": projected_fcf,
            "pv_terminal": round(pv_terminal, 0),
            "base_fcf": round(float(fcf), 0)
        }))
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Candlestick Pattern Recognition ───────────────
@app.route("/api/patterns", methods=["POST"])
def candlestick_patterns():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        prices = data.get("prices", [])
        lookback = int(data.get("lookback_days", 7))
        if len(prices) < 5:
            return safe_jsonify({"patterns": [], "message": "Not enough price data"})

        # Filter to last N trading days (default 7) unless custom range requested
        analysis_prices = prices[-max(lookback + 3, 10):]  # extra context for multi-candle patterns

        predictions = {
            "Doji": "Indecision detected — watch for a breakout in either direction over the next 1-3 days. If followed by a bullish candle, expect upside; if bearish, expect downside.",
            "Hammer": "Buyers are stepping in at lower levels. Expect a potential bullish reversal in the next 2-5 days if volume confirms.",
            "Shooting Star": "Sellers are pushing back at higher levels. Watch for a bearish pullback over the next 2-5 days.",
            "Bullish Engulfing": "Strong buying pressure — expect continuation higher over the next 3-7 days. Consider setting a stop below the engulfing candle low.",
            "Bearish Engulfing": "Strong selling pressure — expect further downside in the next 3-7 days. The engulfing high acts as resistance.",
            "Morning Star": "Classic bottom reversal. Expect a 3-7 day rally from current levels. Pattern reliability: ~70% historically.",
            "Evening Star": "Classic top reversal. Expect a 3-7 day decline. Pattern reliability: ~70% historically.",
            "Bullish Harami": "Selling momentum is fading. Watch for bullish confirmation candle tomorrow — if it appears, expect 2-5 days upside.",
            "Bearish Harami": "Buying momentum is fading. Watch for bearish confirmation candle tomorrow — if it appears, expect 2-5 days downside.",
        }

        patterns_found = []
        for i in range(2, len(analysis_prices)):
            o, h, l, c = analysis_prices[i]["open"], analysis_prices[i]["high"], analysis_prices[i]["low"], analysis_prices[i]["close"]
            body = abs(c - o)
            total_range = h - l
            if total_range == 0: continue
            body_ratio = body / total_range
            prev_o, prev_c = analysis_prices[i-1]["open"], analysis_prices[i-1]["close"]
            prev_body = abs(prev_c - prev_o)
            date = analysis_prices[i]["date"]

            lower_shadow = min(o, c) - l
            upper_shadow = h - max(o, c)

            detected = []

            if body_ratio < 0.1 and total_range > 0:
                detected.append({"pattern": "Doji", "type": "neutral", "description": "Indecision candle — body is tiny relative to range. Potential reversal signal."})

            if body_ratio < 0.35 and lower_shadow > body * 2 and upper_shadow < body * 0.5 and c > o:
                detected.append({"pattern": "Hammer", "type": "bullish", "description": "Bullish reversal — long lower wick shows buyers stepped in."})

            if body_ratio < 0.35 and upper_shadow > body * 2 and lower_shadow < body * 0.5 and c < o:
                detected.append({"pattern": "Shooting Star", "type": "bearish", "description": "Bearish reversal — long upper wick shows sellers pushed price down."})

            if i >= 2 and prev_c < prev_o and c > o and c > prev_o and o < prev_c and body > prev_body:
                detected.append({"pattern": "Bullish Engulfing", "type": "bullish", "description": "Strong bullish reversal — today's green candle engulfs yesterday's red candle."})

            if i >= 2 and prev_c > prev_o and c < o and c < prev_o and o > prev_c and body > prev_body:
                detected.append({"pattern": "Bearish Engulfing", "type": "bearish", "description": "Strong bearish reversal — today's red candle engulfs yesterday's green candle."})

            if i >= 3:
                p2_o, p2_c = analysis_prices[i-2]["open"], analysis_prices[i-2]["close"]
                p1_o, p1_c = analysis_prices[i-1]["open"], analysis_prices[i-1]["close"]
                p2_body = abs(p2_c - p2_o)
                p1_body = abs(p1_c - p1_o)
                if p2_c < p2_o and p2_body > 0 and p1_body < p2_body * 0.3 and c > o and body > p2_body * 0.5:
                    detected.append({"pattern": "Morning Star", "type": "bullish", "description": "Bullish reversal — big red candle, small body, then big green candle."})
                if p2_c > p2_o and p2_body > 0 and p1_body < p2_body * 0.3 and c < o and body > p2_body * 0.5:
                    detected.append({"pattern": "Evening Star", "type": "bearish", "description": "Bearish reversal — big green candle, small body, then big red candle."})

            if i >= 2 and prev_c < prev_o and c > o and o > prev_c and c < prev_o:
                detected.append({"pattern": "Bullish Harami", "type": "bullish", "description": "Potential bullish reversal — small green candle inside previous red candle."})

            if i >= 2 and prev_c > prev_o and c < o and o < prev_c and c > prev_o:
                detected.append({"pattern": "Bearish Harami", "type": "bearish", "description": "Potential bearish reversal — small red candle inside previous green candle."})

            for p in detected:
                p["date"] = date
                p["prediction"] = predictions.get(p["pattern"], "Monitor closely for confirmation.")
                patterns_found.append(p)

        # Aggregate prediction outlook
        bullish = sum(1 for p in patterns_found if p["type"] == "bullish")
        bearish = sum(1 for p in patterns_found if p["type"] == "bearish")
        neutral = sum(1 for p in patterns_found if p["type"] == "neutral")
        if bullish > bearish + neutral:
            outlook = "BULLISH — Majority of recent patterns signal upside. Consider entries on pullbacks."
        elif bearish > bullish + neutral:
            outlook = "BEARISH — Majority of recent patterns signal downside. Exercise caution."
        else:
            outlook = "MIXED/NEUTRAL — Conflicting signals. Wait for clearer pattern confirmation."

        return safe_jsonify({"patterns": patterns_found, "total_found": len(patterns_found),
                            "outlook": outlook, "lookback_days": lookback,
                            "bullish_count": bullish, "bearish_count": bearish, "neutral_count": neutral})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Analyst Ratings & Price Targets (Finnhub) ─────
@app.route("/api/analyst", methods=["POST"])
def analyst_ratings():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        import requests as req

        # Recommendations
        rec_url = f"https://finnhub.io/api/v1/stock/recommendation?symbol={ticker}&token={FINNHUB_KEY}"
        rec_resp = req.get(rec_url, timeout=10).json()
        latest = rec_resp[0] if isinstance(rec_resp, list) and len(rec_resp) > 0 else None

        # Price target
        pt_url = f"https://finnhub.io/api/v1/stock/price-target?symbol={ticker}&token={FINNHUB_KEY}"
        pt_resp = req.get(pt_url, timeout=10).json()

        if latest is None:
            return safe_jsonify({
                "ticker": ticker,
                "recommendation": {"buy": "N/A", "hold": "N/A", "sell": "N/A", "strong_buy": "N/A", "strong_sell": "N/A", "period": "N/A"},
                "price_target": {"high": "N/A", "low": "N/A", "mean": "N/A", "median": "N/A"},
                "available": False
            })

        return safe_jsonify(sanitize_for_json({
            "ticker": ticker,
            "available": True,
            "recommendation": {
                "buy": latest.get("buy", 0), "hold": latest.get("hold", 0),
                "sell": latest.get("sell", 0), "strong_buy": latest.get("strongBuy", 0),
                "strong_sell": latest.get("strongSell", 0), "period": latest.get("period", ""),
            },
            "price_target": {
                "high": pt_resp.get("targetHigh") if pt_resp.get("targetHigh") else "N/A",
                "low": pt_resp.get("targetLow") if pt_resp.get("targetLow") else "N/A",
                "mean": pt_resp.get("targetMean") if pt_resp.get("targetMean") else "N/A",
                "median": pt_resp.get("targetMedian") if pt_resp.get("targetMedian") else "N/A",
            }
        }))
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Insider Trading (Finnhub) ─────────────────────
@app.route("/api/insider", methods=["POST"])
def insider_trading():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        import requests as req
        url = f"https://finnhub.io/api/v1/stock/insider-transactions?symbol={ticker}&token={FINNHUB_KEY}"
        resp = req.get(url, timeout=10).json()
        txns = resp.get("data", [])[:20]  # Last 20

        result = []
        for t in txns:
            result.append({
                "name": t.get("name", "Unknown"),
                "share": safe_num(t.get("share", 0), 0),
                "change": safe_num(t.get("change", 0), 0),
                "transaction_type": t.get("transactionType", ""),
                "filing_date": t.get("filingDate", ""),
                "transaction_date": t.get("transactionDate", ""),
            })

        return safe_jsonify({"ticker": ticker, "transactions": result})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Market Sentiment (Fear & Greed + VIX) ─────────
@app.route("/api/sentiment-market", methods=["POST"])
def market_sentiment():
    try:
        result = {}
        # Fear & Greed Index
        try:
            import fear_and_greed
            fg = fear_and_greed.get()
            result["fear_greed"] = {
                "value": round(fg.value, 1),
                "description": fg.description,
            }
        except Exception as e:
            result["fear_greed"] = {"value": 50, "description": "Neutral", "error": str(e)}

        # VIX
        try:
            import yfinance as yf
            vix = yf.Ticker("^VIX")
            vix_hist = vix.history(period="5d")
            if not vix_hist.empty:
                vix_close = float(vix_hist["Close"].iloc[-1])
                result["vix"] = {
                    "value": round(vix_close, 2),
                    "label": "Low Volatility" if vix_close < 15 else "Moderate" if vix_close < 25 else "High Volatility" if vix_close < 35 else "Extreme Fear"
                }
            else:
                result["vix"] = {"value": "N/A", "label": "Unavailable"}
        except Exception as e:
            result["vix"] = {"value": "N/A", "label": "Unavailable", "error": str(e)}

        return safe_jsonify(result)
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

# ── Route: FRED Economic Indicators ──────────────────────
@app.route("/api/macro", methods=["POST"])
def macro_indicators():
    try:
        from fredapi import Fred
        fred = Fred(api_key=FRED_KEY)
        indicators = {}
        series_map = {
            "gdp_growth": ("GDPC1", "GDP (Real, Quarterly)"),
            "cpi": ("CPIAUCSL", "Consumer Price Index"),
            "fed_rate": ("FEDFUNDS", "Federal Funds Rate"),
            "unemployment": ("UNRATE", "Unemployment Rate"),
            "treasury_2y": ("DGS2", "2-Year Treasury Yield"),
            "treasury_10y": ("DGS10", "10-Year Treasury Yield"),
            "treasury_30y": ("DGS30", "30-Year Treasury Yield"),
        }

        def assess_indicator(key, val):
            """Returns (status, explanation) — status is 'good', 'bad', or 'neutral'"""
            if val == "N/A": return "neutral", "Data unavailable."
            v = float(val)
            assessments = {
                "gdp_growth": lambda v: ("good", f"GDP at ${v:,.0f}B indicates a strong economy. Higher GDP means more corporate earnings.") if v > 20000 else ("neutral", f"GDP at ${v:,.0f}B — moderate economic output."),
                "cpi": lambda v: ("bad", f"CPI at {v:.1f} — inflation is elevated, eroding purchasing power and pressuring the Fed to keep rates high.") if v > 310 else ("good", f"CPI at {v:.1f} — inflation is relatively contained, supportive of lower interest rates."),
                "fed_rate": lambda v: ("good", f"Fed rate at {v:.2f}% — low rates stimulate borrowing and boost stock valuations.") if v < 2.5 else (("neutral", f"Fed rate at {v:.2f}% — moderate rates balance growth and inflation.") if v < 4.5 else ("bad", f"Fed rate at {v:.2f}% — high rates increase borrowing costs and can weigh on equity valuations.")),
                "unemployment": lambda v: ("good", f"Unemployment at {v:.1f}% — tight labor market indicates strong economic health.") if v < 4.0 else (("neutral", f"Unemployment at {v:.1f}% — within normal range but rising unemployment could signal slowing growth.") if v < 5.5 else ("bad", f"Unemployment at {v:.1f}% — elevated joblessness signals economic weakness.")),
                "treasury_2y": lambda v: ("neutral", f"2Y yield at {v:.2f}% — reflects market's short-term rate expectations. Higher yield means tighter policy expected.") if v < 4.5 else ("bad", f"2Y yield at {v:.2f}% — elevated short-term yields signal aggressive monetary tightening."),
                "treasury_10y": lambda v: ("good", f"10Y yield at {v:.2f}% — low long-term rates are supportive of equity valuations and reduce borrowing costs.") if v < 3.5 else (("neutral", f"10Y yield at {v:.2f}% — moderate levels. Stocks can handle this but it raises cost of capital.") if v < 4.5 else ("bad", f"10Y yield at {v:.2f}% — high yields compete with stocks for investor capital and increase discount rates.")),
                "treasury_30y": lambda v: ("neutral", f"30Y yield at {v:.2f}% — reflects long-term growth and inflation expectations.") if v < 4.0 else ("bad", f"30Y yield at {v:.2f}% — elevated long-term yields suggest persistent inflation concerns."),
            }
            fn = assessments.get(key)
            if fn: return fn(v)
            return "neutral", "No assessment available."

        for key, (series_id, label) in series_map.items():
            try:
                s = fred.get_series(series_id)
                latest = s.dropna().iloc[-1]
                val = round(float(latest), 2)
                status, explanation = assess_indicator(key, val)
                indicators[key] = {"value": val, "label": label, "status": status, "explanation": explanation}
            except:
                indicators[key] = {"value": "N/A", "label": label, "status": "neutral", "explanation": "Data currently unavailable."}

        return safe_jsonify({"indicators": indicators})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Earnings Calendar (Finnhub) ───────────────────
@app.route("/api/earnings", methods=["POST"])
def earnings_calendar():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        import requests as req
        url = f"https://finnhub.io/api/v1/stock/earnings?symbol={ticker}&token={FINNHUB_KEY}"
        resp = req.get(url, timeout=10).json()

        earnings = []
        for e in resp[:12]:  # Last 12 quarters
            earnings.append(sanitize_for_json({
                "period": e.get("period", ""),
                "actual": e.get("actual", "N/A"),
                "estimate": e.get("estimate", "N/A"),
                "surprise": e.get("surprise", "N/A"),
                "surprise_pct": e.get("surprisePercent", "N/A"),
            }))

        return safe_jsonify({"ticker": ticker, "earnings": earnings})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Monte Carlo Simulation ────────────────────────
@app.route("/api/monte-carlo", methods=["POST"])
def monte_carlo():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        days = int(data.get("days", 60))
        simulations = int(data.get("simulations", 1000))
        prices = data.get("prices", [])

        if len(prices) < 30:
            return safe_jsonify({"error": "Need at least 30 data points"})

        closes = np.array([p["close"] for p in prices], dtype=float)
        returns = np.diff(np.log(closes))
        mu = np.mean(returns)
        sigma = np.std(returns)
        last_price = closes[-1]

        # Run simulations
        np.random.seed(42)
        all_paths = np.zeros((simulations, days))
        for i in range(simulations):
            daily_returns = np.random.normal(mu, sigma, days)
            price_path = last_price * np.exp(np.cumsum(daily_returns))
            all_paths[i] = price_path

        # Calculate percentile bands
        percentiles = {}
        for p in [10, 25, 50, 75, 90]:
            band = np.percentile(all_paths, p, axis=0)
            percentiles[f"p{p}"] = [round(float(v), 2) for v in band]

        final_prices = all_paths[:, -1]
        return safe_jsonify(sanitize_for_json({
            "ticker": ticker, "days": days, "simulations": simulations,
            "start_price": round(float(last_price), 2),
            "percentiles": percentiles,
            "final_stats": {
                "mean": round(float(np.mean(final_prices)), 2),
                "median": round(float(np.median(final_prices)), 2),
                "std": round(float(np.std(final_prices)), 2),
                "p10": round(float(np.percentile(final_prices, 10)), 2),
                "p90": round(float(np.percentile(final_prices, 90)), 2),
            }
        }))
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Altman Z-Score ────────────────────────────────
@app.route("/api/zscore", methods=["POST"])
def altman_zscore():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        tk = Toolkit(tickers=[ticker], api_key=FMP_API_KEY)
        balance = tk.get_balance_sheet_statement()
        income = tk.get_income_statement()

        ta = get_fin_val(balance, ticker, "Total Assets")
        tl = get_fin_val(balance, ticker, "Total Liabilities")
        ca = get_fin_val(balance, ticker, "Total Current Assets")
        cl = get_fin_val(balance, ticker, "Total Current Liabilities")
        re = get_fin_val(balance, ticker, "Retained Earnings")
        ebit = get_fin_val(income, ticker, "Operating Income")
        revenue = get_fin_val(income, ticker, "Revenue")
        te = get_fin_val(balance, ticker, "Total Equity")

        if any(v == "N/A" or v == 0 for v in [ta]):
            return safe_jsonify({"error": "Insufficient data", "ticker": ticker})

        ta, tl = float(ta), float(tl) if tl != "N/A" else 0
        ca = float(ca) if ca != "N/A" else 0
        cl = float(cl) if cl != "N/A" else 0
        re = float(re) if re != "N/A" else 0
        ebit = float(ebit) if ebit != "N/A" else 0
        revenue = float(revenue) if revenue != "N/A" else 0
        te = float(te) if te != "N/A" else 0

        # Get market cap for ratio D
        import yfinance as yf
        stock = yf.Ticker(ticker)
        market_cap = stock.info.get("marketCap", 0) or 0

        wc = ca - cl
        A = wc / ta if ta != 0 else 0
        B = re / ta if ta != 0 else 0
        C = ebit / ta if ta != 0 else 0
        D = market_cap / tl if tl != 0 else 0
        E = revenue / ta if ta != 0 else 0

        z = 1.2 * A + 1.4 * B + 3.3 * C + 0.6 * D + E

        if z > 2.99: zone = "Safe Zone"
        elif z > 1.81: zone = "Grey Zone"
        else: zone = "Distress Zone"

        return safe_jsonify(sanitize_for_json({
            "ticker": ticker, "z_score": round(z, 2), "zone": zone,
            "components": {
                "A_working_capital": round(A, 4), "B_retained_earnings": round(B, 4),
                "C_ebit": round(C, 4), "D_market_cap_debt": round(D, 4),
                "E_revenue": round(E, 4),
            }
        }))
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Dividend Analysis ─────────────────────────────
@app.route("/api/dividends", methods=["POST"])
def dividend_analysis():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        import yfinance as yf
        stock = yf.Ticker(ticker)
        info = stock.info
        dividends = stock.dividends

        div_history = []
        if dividends is not None and len(dividends) > 0:
            for date, amount in dividends.tail(20).items():
                div_history.append({"date": str(date)[:10], "amount": round(float(amount), 4)})

        return safe_jsonify(sanitize_for_json({
            "ticker": ticker,
            "dividend_yield": info.get("dividendYield", "N/A"),
            "dividend_rate": info.get("dividendRate", "N/A"),
            "payout_ratio": info.get("payoutRatio", "N/A"),
            "ex_dividend_date": str(info.get("exDividendDate", "N/A")),
            "five_year_avg_yield": info.get("fiveYearAvgDividendYield", "N/A"),
            "history": div_history,
        }))
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Correlation Matrix ────────────────────────────
@app.route("/api/correlation", methods=["POST"])
def correlation_matrix():
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        import yfinance as yf

        benchmarks = ["SPY", "QQQ", "DIA", "GLD", "TLT", "VIX"]
        all_tickers = [ticker] + [b for b in benchmarks if b.upper() != ticker]

        closes = {}
        for t in all_tickers:
            try:
                sym = f"^{t}" if t == "VIX" else t
                hist = yf.Ticker(sym).history(period="1y")
                if not hist.empty:
                    closes[t] = hist["Close"].pct_change().dropna()
            except: pass

        if len(closes) < 2:
            return safe_jsonify({"error": "Not enough data for correlation"})

        df = pd.DataFrame(closes)
        corr = df.corr()

        matrix = {}
        for col in corr.columns:
            matrix[col] = {}
            for row in corr.index:
                matrix[col][row] = round(float(corr.loc[row, col]), 3)

        return safe_jsonify({"ticker": ticker, "tickers": list(corr.columns), "matrix": matrix})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Sector Performance Heatmap (Alpha Vantage) ────
@app.route("/api/heatmap", methods=["POST"])
def sector_heatmap():
    try:
        import requests as req
        url = f"https://www.alphavantage.co/query?function=SECTOR&apikey={ALPHA_VANTAGE_KEY}"
        resp = req.get(url, timeout=15).json()

        sectors = {}
        timeframes = {
            "1D": "Rank A: Real-Time Performance",
            "1W": "Rank C: 5 Day Performance",
            "1M": "Rank D: 1 Month Performance",
            "3M": "Rank E: 3 Month Performance",
            "YTD": "Rank F: Year-to-Date (YTD) Performance",
            "1Y": "Rank G: 1 Year Performance",
        }

        for tf_key, av_key in timeframes.items():
            data = resp.get(av_key, {})
            sectors[tf_key] = {}
            for sector, pct in data.items():
                try:
                    sectors[tf_key][sector] = float(pct.replace("%", ""))
                except:
                    sectors[tf_key][sector] = 0

        return safe_jsonify({"sectors": sectors, "timeframes": list(timeframes.keys())})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Currency Conversion Rates ─────────────────────
@app.route("/api/currency", methods=["POST"])
def currency_rates():
    """Return conversion rates from USD to GBP/INR so the frontend can convert all values."""
    try:
        import requests as req
        url = f"https://www.alphavantage.co/query?function=CURRENCY_EXCHANGE_RATE&from_currency=USD&to_currency=GBP&apikey={ALPHA_VANTAGE_KEY}"
        resp = req.get(url, timeout=10).json()
        usd_gbp = float(resp.get("Realtime Currency Exchange Rate", {}).get("5. Exchange Rate", 0.79))
    except:
        usd_gbp = 0.79  # Fallback
    try:
        import requests as req
        url = f"https://www.alphavantage.co/query?function=CURRENCY_EXCHANGE_RATE&from_currency=USD&to_currency=INR&apikey={ALPHA_VANTAGE_KEY}"
        resp = req.get(url, timeout=10).json()
        usd_inr = float(resp.get("Realtime Currency Exchange Rate", {}).get("5. Exchange Rate", 83.5))
    except:
        usd_inr = 83.5
    return safe_jsonify({"USD": 1.0, "GBP": usd_gbp, "INR": usd_inr})

# ── Route: Options Chain (with Greeks) ───────────────────
@app.route("/api/options/chain", methods=["POST"])
def options_chain():
    """Fetch options chain data with calculated Greeks."""
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        import yfinance as yf
        from options_engine import enrich_chain_with_greeks
        stock = yf.Ticker(ticker)
        expirations = stock.options
        if not expirations:
            return safe_jsonify({"error": f"No options data available for {ticker}", "available": False})

        exp = data.get("expiration", expirations[0])
        if exp not in expirations:
            exp = expirations[0]

        chain = stock.option_chain(exp)
        calls = chain.calls.fillna(0).to_dict(orient="records")
        puts = chain.puts.fillna(0).to_dict(orient="records")

        # Clean up timestamps
        for row in calls + puts:
            for k, v in row.items():
                if hasattr(v, 'isoformat'):
                    row[k] = str(v)

        info = stock.info
        current_price = info.get("regularMarketPrice") or info.get("previousClose", 0)

        # Enrich with Greeks
        for c in calls: c["_type"] = "call"
        for p in puts: p["_type"] = "put"
        calls = enrich_chain_with_greeks(calls, current_price, exp)
        puts = enrich_chain_with_greeks(puts, current_price, exp)

        return safe_jsonify({
            "ticker": ticker,
            "current_price": current_price,
            "expiration": exp,
            "expirations": list(expirations),
            "calls": calls[:50],
            "puts": puts[:50],
            "call_count": len(calls),
            "put_count": len(puts),
        })
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Futures Data ──────────────────────────────────
@app.route("/api/futures", methods=["POST"])
def futures_data():
    """Fetch futures contract data via yfinance."""
    try:
        data = request.json
        ticker = data.get("ticker", "ES=F").upper().strip()
        import yfinance as yf
        fut = yf.Ticker(ticker)
        info = fut.info

        # Historical data
        hist = fut.history(period="6mo")
        price_history = []
        for date, row in hist.iterrows():
            price_history.append({
                "date": date.strftime("%Y-%m-%d"),
                "open": round(row["Open"], 2),
                "high": round(row["High"], 2),
                "low": round(row["Low"], 2),
                "close": round(row["Close"], 2),
                "volume": int(row.get("Volume", 0)),
            })

        return safe_jsonify({
            "ticker": ticker,
            "name": info.get("shortName", ticker),
            "price": info.get("regularMarketPrice") or info.get("previousClose", 0),
            "change": info.get("regularMarketChange", 0),
            "change_pct": info.get("regularMarketChangePercent", 0),
            "day_high": info.get("regularMarketDayHigh", 0),
            "day_low": info.get("regularMarketDayLow", 0),
            "open_interest": info.get("openInterest", "N/A"),
            "volume": info.get("regularMarketVolume", 0),
            "prev_close": info.get("regularMarketPreviousClose", 0),
            "price_history": price_history,
        })
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Forex Pair Data ───────────────────────────────
@app.route("/api/forex", methods=["POST"])
def forex_data():
    """Fetch forex pair data via yfinance."""
    try:
        data = request.json
        pair = data.get("pair", "EURUSD=X").upper().strip()
        import yfinance as yf
        fx = yf.Ticker(pair)
        info = fx.info

        hist = fx.history(period="6mo")
        price_history = []
        for date, row in hist.iterrows():
            price_history.append({
                "date": date.strftime("%Y-%m-%d"),
                "open": round(row["Open"], 6),
                "high": round(row["High"], 6),
                "low": round(row["Low"], 6),
                "close": round(row["Close"], 6),
                "volume": int(row.get("Volume", 0)),
            })

        return safe_jsonify({
            "pair": pair,
            "name": info.get("shortName", pair),
            "rate": info.get("regularMarketPrice") or info.get("previousClose", 0),
            "change": info.get("regularMarketChange", 0),
            "change_pct": info.get("regularMarketChangePercent", 0),
            "day_high": info.get("regularMarketDayHigh", 0),
            "day_low": info.get("regularMarketDayLow", 0),
            "bid": info.get("bid", 0),
            "ask": info.get("ask", 0),
            "price_history": price_history,
        })
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Live Price (for auto-refresh) ─────────────────
@app.route("/api/live-price", methods=["POST"])
def live_price():
    """Quick price fetch for auto-refresh polling."""
    try:
        data = request.json
        ticker = data.get("ticker", "AAPL").upper().strip()
        import yfinance as yf
        stock = yf.Ticker(ticker)
        info = stock.info
        return safe_jsonify({
            "ticker": ticker,
            "price": info.get("regularMarketPrice") or info.get("previousClose", 0),
            "change": info.get("regularMarketChange", 0),
            "change_pct": info.get("regularMarketChangePercent", 0),
            "volume": info.get("regularMarketVolume", 0),
            "timestamp": datetime.now().isoformat(),
        })
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Options Payoff Diagram ─────────────────────
@app.route("/api/options/payoff", methods=["POST"])
def options_payoff():
    try:
        from options_engine import payoff_diagram
        data = request.json
        K = float(data.get("strike", 100))
        premium = float(data.get("premium", 5))
        opt_type = data.get("option_type", "call")
        is_long = data.get("is_long", True)
        S = float(data.get("current_price", K))
        points = payoff_diagram(S, K, premium, opt_type, is_long)
        return safe_jsonify({"points": points, "strike": K, "premium": premium, "type": opt_type, "direction": "long" if is_long else "short"})
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Portfolio ──────────────────────────────────
import portfolio_engine

def _fetch_live_prices(tickers):
    """Helper: fetch live prices for a list of tickers."""
    import yfinance as yf
    prices = {}
    for t in tickers:
        try:
            stock = yf.Ticker(t)
            info = stock.info
            prices[t] = info.get("regularMarketPrice") or info.get("previousClose", 0)
        except:
            pass
    return prices

@app.route("/api/portfolio", methods=["GET"])
def get_portfolio():
    """Get current portfolio with live prices."""
    try:
        positions = portfolio_engine.get_positions()
        current_prices = _fetch_live_prices([p["ticker"] for p in positions])
        summary = portfolio_engine.get_portfolio_summary(current_prices)
        # Record equity point and daily snapshot
        portfolio_engine.record_equity_point(summary["total_value"], summary["cash"], summary["positions_value"])
        portfolio_engine.save_daily_snapshot(summary["total_value"], summary["cash"], summary["positions_value"])
        # Check pending orders
        filled = portfolio_engine.check_and_fill_orders(current_prices)
        summary["filled_orders"] = filled
        summary["pending_orders"] = portfolio_engine.get_pending_orders()
        return safe_jsonify(summary)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/buy", methods=["POST"])
def portfolio_buy():
    try:
        data = request.json
        ticker = data.get("ticker", "").upper().strip()
        shares = float(data.get("shares", 0))
        asset_type = data.get("asset_type", "stock")
        import yfinance as yf
        stock = yf.Ticker(ticker)
        info = stock.info
        price = info.get("regularMarketPrice") or info.get("previousClose", 0)
        if price <= 0:
            return safe_jsonify({"error": f"Cannot get price for {ticker}"}), 400
        result = portfolio_engine.buy(ticker, shares, price, asset_type)
        if "error" in result:
            return safe_jsonify(result), 400
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/sell", methods=["POST"])
def portfolio_sell():
    try:
        data = request.json
        ticker = data.get("ticker", "").upper().strip()
        shares = float(data.get("shares", 0))
        import yfinance as yf
        stock = yf.Ticker(ticker)
        info = stock.info
        price = info.get("regularMarketPrice") or info.get("previousClose", 0)
        if price <= 0:
            return safe_jsonify({"error": f"Cannot get price for {ticker}"}), 400
        result = portfolio_engine.sell(ticker, shares, price)
        if "error" in result:
            return safe_jsonify(result), 400
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/short", methods=["POST"])
def portfolio_short():
    """Open a short position."""
    try:
        data = request.json
        ticker = data.get("ticker", "").upper().strip()
        shares = float(data.get("shares", 0))
        asset_type = data.get("asset_type", "stock")
        import yfinance as yf
        stock = yf.Ticker(ticker)
        info = stock.info
        price = info.get("regularMarketPrice") or info.get("previousClose", 0)
        if price <= 0:
            return safe_jsonify({"error": f"Cannot get price for {ticker}"}), 400
        result = portfolio_engine.short_sell(ticker, shares, price, asset_type)
        if "error" in result:
            return safe_jsonify(result), 400
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/cover", methods=["POST"])
def portfolio_cover():
    """Cover (close) a short position."""
    try:
        data = request.json
        ticker = data.get("ticker", "").upper().strip()
        shares = float(data.get("shares", 0))
        import yfinance as yf
        stock = yf.Ticker(ticker)
        info = stock.info
        price = info.get("regularMarketPrice") or info.get("previousClose", 0)
        if price <= 0:
            return safe_jsonify({"error": f"Cannot get price for {ticker}"}), 400
        result = portfolio_engine.cover_short(ticker, shares, price)
        if "error" in result:
            return safe_jsonify(result), 400
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/limit-order", methods=["POST"])
def portfolio_limit_order():
    """Place a limit order."""
    try:
        data = request.json
        result = portfolio_engine.place_limit_order(
            ticker=data.get("ticker", "").upper().strip(),
            side=data.get("side", "BUY").upper(),
            shares=float(data.get("shares", 0)),
            limit_price=float(data.get("price", 0)),
            asset_type=data.get("asset_type", "stock"),
        )
        if "error" in result:
            return safe_jsonify(result), 400
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/stop-order", methods=["POST"])
def portfolio_stop_order():
    """Place a stop order (stop-loss / take-profit)."""
    try:
        data = request.json
        result = portfolio_engine.place_stop_order(
            ticker=data.get("ticker", "").upper().strip(),
            side=data.get("side", "SELL").upper(),
            shares=float(data.get("shares", 0)),
            stop_price=float(data.get("price", 0)),
            asset_type=data.get("asset_type", "stock"),
        )
        if "error" in result:
            return safe_jsonify(result), 400
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/orders", methods=["GET"])
def portfolio_orders():
    """Get pending orders."""
    try:
        return safe_jsonify({"orders": portfolio_engine.get_pending_orders()})
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/cancel-order", methods=["POST"])
def portfolio_cancel_order():
    """Cancel a pending order."""
    try:
        data = request.json
        order_id = int(data.get("order_id", 0))
        return safe_jsonify(portfolio_engine.cancel_order(order_id))
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/equity-curve", methods=["GET"])
def portfolio_equity_curve():
    """Get equity curve for charting."""
    try:
        return safe_jsonify({"curve": portfolio_engine.get_equity_curve()})
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/history", methods=["GET"])
def portfolio_history():
    try:
        return safe_jsonify({"transactions": portfolio_engine.get_transactions()})
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/analytics", methods=["GET"])
def portfolio_analytics():
    try:
        positions = portfolio_engine.get_positions()
        current_prices = _fetch_live_prices([p["ticker"] for p in positions])
        return safe_jsonify(portfolio_engine.get_analytics(current_prices))
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

@app.route("/api/portfolio/reset", methods=["POST"])
def portfolio_reset():
    try:
        return safe_jsonify(portfolio_engine.reset_portfolio())
    except Exception as e:
        return safe_jsonify({"error": str(e)}), 500

# ── Route: Market Summary (Dashboard) ────────────────
_market_cache = {"data": None, "ts": 0}

@app.route("/api/market-summary", methods=["GET"])
def market_summary():
    """Global market snapshot for the dashboard landing page."""
    import time
    now = time.time()
    # 5-minute cache
    if _market_cache["data"] and (now - _market_cache["ts"]) < 300:
        return safe_jsonify(_market_cache["data"])

    try:
        import yfinance as yf

        symbols = {
            "indices": [
                ("^NSEI", "NIFTY 50"), ("^BSESN", "SENSEX"),
                ("^GSPC", "S&P 500"), ("^IXIC", "NASDAQ"),
                ("^FTSE", "FTSE 100"), ("^N225", "Nikkei 225"),
            ],
            "commodities": [
                ("GC=F", "Gold"), ("SI=F", "Silver"),
                ("CL=F", "Crude Oil WTI"), ("NG=F", "Natural Gas"),
            ],
            "currencies": [
                ("EURUSD=X", "EUR/USD"), ("GBPUSD=X", "GBP/USD"),
                ("USDJPY=X", "USD/JPY"), ("USDINR=X", "USD/INR"),
            ],
        }

        result = {}
        for category, items in symbols.items():
            cat_list = []
            tickers_str = " ".join([s[0] for s in items])
            try:
                data = yf.download(tickers_str, period="2d", group_by="ticker", progress=False)
                for sym, name in items:
                    try:
                        if len(items) == 1:
                            df = data
                        else:
                            df = data[sym] if sym in data.columns.get_level_values(0) else None

                        if df is not None and len(df) >= 1:
                            latest = df.iloc[-1]
                            prev = df.iloc[-2] if len(df) >= 2 else df.iloc[-1]
                            price = float(latest["Close"])
                            prev_close = float(prev["Close"])
                            change = price - prev_close
                            change_pct = (change / prev_close * 100) if prev_close else 0
                            cat_list.append({
                                "symbol": sym, "name": name,
                                "price": round(price, 2),
                                "change": round(change, 2),
                                "change_pct": round(change_pct, 2),
                            })
                        else:
                            cat_list.append({"symbol": sym, "name": name, "price": 0, "change": 0, "change_pct": 0})
                    except Exception:
                        cat_list.append({"symbol": sym, "name": name, "price": 0, "change": 0, "change_pct": 0})
            except Exception:
                for sym, name in items:
                    cat_list.append({"symbol": sym, "name": name, "price": 0, "change": 0, "change_pct": 0})
            result[category] = cat_list

        _market_cache["data"] = result
        _market_cache["ts"] = now
        return safe_jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return safe_jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    print(f"\n  FinanceIQ v5.2")
    print(f"  Local:   http://localhost:5000")
    print(f"  Network: http://{local_ip}:5000\n")
    app.run(debug=True, port=5000, host="0.0.0.0")

